"""
国铁电子客票 Web 解析工具 - Flask 后端
运行: python app.py
访问: http://localhost:5678

核心解析逻辑已抽取到根目录 parser.py 共用。
"""

import io
import csv
import os
from flask import Flask, request, jsonify, send_file, render_template

from parser import (
    FIELDNAMES,
    process_pdf_bytes,
    extract_text_from_bytes,
    parse_guotie,
)


app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 100 * 1024 * 1024  # 100MB 上限


# ── 路由 ──────────────────────────────────────────────────
@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/parse", methods=["POST"])
def api_parse():
    """批量解析接口（一次性上传所有文件）。"""
    files = request.files.getlist("files")
    if not files:
        return jsonify({"error": "未收到文件"}), 400

    records = []
    for f in files:
        if not f.filename.lower().endswith(".pdf"):
            continue
        pdf_bytes = f.read()
        rec = process_pdf_bytes(pdf_bytes, f.filename)
        records.append(rec)

    return jsonify({"records": records, "fields": FIELDNAMES})


@app.route("/api/parse_one", methods=["POST"])
def api_parse_one():
    """逐文件解析接口（流式进度用）。"""
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "未收到文件"}), 400
    pdf_bytes = f.read()
    rec = process_pdf_bytes(pdf_bytes, f.filename)
    return jsonify({"record": rec, "fields": FIELDNAMES})


@app.route("/api/download_csv", methods=["POST"])
def api_download_csv():
    """下载 CSV。"""
    data = request.get_json()
    records = data.get("records", [])

    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=FIELDNAMES, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(records)

    csv_bytes = io.BytesIO(("\ufeff" + buf.getvalue()).encode("utf-8"))
    return send_file(
        csv_bytes,
        mimetype="text/csv; charset=utf-8",
        as_attachment=True,
        download_name="tickets.csv",
    )


@app.route("/api/download_excel", methods=["POST"])
def api_download_excel():
    """下载 Excel（.xlsx）。"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({"error": "服务端未安装 openpyxl，请联系管理员"}), 500

    data = request.get_json()
    records = data.get("records", [])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "国铁发票"

    # 标题行样式
    header_fill = PatternFill("solid", fgColor="1677FF")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin = Side(style="thin", color="D9D9D9")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(FIELDNAMES)
    for col_idx, _ in enumerate(FIELDNAMES, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill  = header_fill
        cell.font  = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = cell_border

    # 数据行
    for row_idx, rec in enumerate(records, 2):
        for col_idx, field in enumerate(FIELDNAMES, 1):
            val = rec.get(field, "")
            # 票价尝试转数字
            if field == "票价(元)" and val:
                try:
                    val = float(val)
                except ValueError:
                    pass
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = cell_border
            cell.alignment = Alignment(vertical="center")
            # 奇偶行底色
            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F5F8FF")

    # 自适应列宽
    for col_idx, field in enumerate(FIELDNAMES, 1):
        max_len = len(field)
        for row_idx in range(2, len(records) + 2):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="tickets.xlsx",
    )


# 兼容旧接口路径
@app.route("/api/download", methods=["POST"])
def api_download():
    return api_download_csv()


if __name__ == "__main__":
    app.run(debug=False, port=5678)