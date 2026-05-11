"""
增值税电子普通发票 Web 解析工具 - Flask 后端
运行: python app.py
访问: http://localhost:5678

核心解析逻辑已抽取到根目录 fapiao_parser.py 共用。
"""

import io
import csv
import os
from collections import defaultdict
from flask import Flask, request, jsonify, send_file, render_template

import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from fapiao_parser import (
    FIELDNAMES,
    process_pdf_bytes,
    extract_text_from_bytes,
    parse_fapiao,
)

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 100 * 1024 * 1024  # 100MB 上限

DATA_FILE = os.path.join(os.path.dirname(__file__), "static", "invoice_data.xlsx")


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/parse", methods=["POST"])
def api_parse():
    """批量解析接口（一次性上传所有文件）。"""
    files = request.files.getlist("files")
    if not files:
        return jsonify({"error": "未收到文件"}), 400

    records = []
    for f in files:
        if not f.filename.lower().endswith(".pdf"):
            continue
        pdf_bytes = f.read()
        rec = process_pdf_bytes(pdf_bytes, f.filename)
        records.append(rec)

    return jsonify({"records": records, "fields": FIELDNAMES})


@app.route("/api/parse_one", methods=["POST"])
def api_parse_one():
    """逐文件解析接口（流式进度用）。"""
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "未收到文件"}), 400
    pdf_bytes = f.read()
    rec = process_pdf_bytes(pdf_bytes, f.filename)
    return jsonify({"record": rec, "fields": FIELDNAMES})


@app.route("/api/load_data", methods=["GET"])
def api_load_data():
    """加载本地Excel数据。"""
    if not os.path.exists(DATA_FILE):
        return jsonify({"records": [], "fields": FIELDNAMES})

    try:
        import openpyxl
        wb = openpyxl.load_workbook(DATA_FILE)
        ws = wb.active
        if ws.max_row < 2:
            return jsonify({"records": [], "fields": FIELDNAMES})

        headers = [cell.value for cell in ws[1]]
        records = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            rec = {}
            for i, h in enumerate(headers):
                if h and i < len(row):
                    rec[h] = row[i] if row[i] is not None else ""
            if rec:
                for fn in FIELDNAMES:
                    if fn not in rec:
                        rec[fn] = ""
                records.append(rec)

        wb.close()
        return jsonify({"records": records, "fields": FIELDNAMES})
    except Exception as e:
        return jsonify({"error": f"加载数据失败: {e}"}), 500


@app.route("/api/save_data", methods=["POST"])
def api_save_data():
    """保存数据到本地Excel。"""
    data = request.get_json()
    records = data.get("records", [])

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        os.makedirs(os.path.dirname(DATA_FILE), exist_ok=True)

        wb = openpyxl.Workbook()
        header_fill = PatternFill("solid", fgColor="1677FF")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        thin = Side(style="thin", color="D9D9D9")
        cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws = wb.active
        ws.title = "发票明细"

        ws.append(FIELDNAMES)
        for col_idx, _ in enumerate(FIELDNAMES, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = cell_border

        for row_idx, rec in enumerate(records, 2):
            for col_idx, field in enumerate(FIELDNAMES, 1):
                val = rec.get(field, "")
                if field in ["金额", "税额", "价税合计", "单价", "数量"] and val:
                    try:
                        val = float(val)
                    except (ValueError, TypeError):
                        val = val
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = cell_border
                cell.alignment = Alignment(vertical="center")
                if row_idx % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor="F5F8FF")

        from openpyxl.utils import get_column_letter
        for col_idx, field in enumerate(FIELDNAMES, 1):
            max_len = len(field)
            for row_idx in range(2, len(records) + 2):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 28

        wb.save(DATA_FILE)
        wb.close()

        return jsonify({"success": True, "message": f"已保存 {len(records)} 条记录"})
    except Exception as e:
        return jsonify({"error": f"保存失败: {e}"}), 500


@app.route("/api/download_csv", methods=["POST"])
def api_download_csv():
    """下载 CSV。"""
    data = request.get_json()
    records = data.get("records", [])

    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=FIELDNAMES, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(records)

    csv_bytes = io.BytesIO(("\ufeff" + buf.getvalue()).encode("utf-8"))
    return send_file(
        csv_bytes,
        mimetype="text/csv; charset=utf-8",
        as_attachment=True,
        download_name="fapiao.csv",
    )


@app.route("/api/download_excel", methods=["POST"])
def api_download_excel():
    """下载 Excel（.xlsx）- 多Sheet结构。"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({"error": "服务端未安装 openpyxl，请联系管理员"}), 500

    data = request.get_json()
    records = data.get("records", [])

    wb = openpyxl.Workbook()

    header_fill = PatternFill("solid", fgColor="1677FF")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    title_font = Font(bold=True, size=14)
    thin = Side(style="thin", color="D9D9D9")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws1 = wb.active
    ws1.title = "发票明细"
    _write_detail_sheet(ws1, records, header_fill, header_font, cell_border)

    ws2 = wb.create_sheet("月度统计")
    _write_month_stats_sheet(ws2, records, header_fill, header_font, cell_border, title_font)

    ws3 = wb.create_sheet("销售方分析")
    _write_seller_stats_sheet(ws3, records, header_fill, header_font, cell_border, title_font)

    ws4 = wb.create_sheet("报销汇总")
    _write_reimb_stats_sheet(ws4, records, header_fill, header_font, cell_border, title_font)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="fapiao.xlsx",
    )


def _write_detail_sheet(ws, records, header_fill, header_font, cell_border):
    from openpyxl.styles import Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    ws.append(FIELDNAMES)
    for col_idx, _ in enumerate(FIELDNAMES, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = cell_border

    for row_idx, rec in enumerate(records, 2):
        for col_idx, field in enumerate(FIELDNAMES, 1):
            val = rec.get(field, "")
            if field in ["金额", "税额", "价税合计", "单价", "数量"] and val:
                try:
                    val = float(val)
                except (ValueError, TypeError):
                    pass
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = cell_border
            cell.alignment = Alignment(vertical="center")
            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F5F8FF")

    for col_idx, field in enumerate(FIELDNAMES, 1):
        max_len = len(field)
        for row_idx in range(2, len(records) + 2):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28


def _write_month_stats_sheet(ws, records, header_fill, header_font, cell_border, title_font):
    from openpyxl.styles import Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    month_map = defaultdict(lambda: {"count": 0, "amount": 0.0})

    for rec in records:
        date_str = rec.get("开票日期", "")
        amount = _parse_amount(rec.get("价税合计", ""))
        m = _parse_date(date_str)
        if m:
            month_key = f"{m['year']}-{str(m['month']).zfill(2)}"
            month_map[month_key]["count"] += 1
            month_map[month_key]["amount"] += amount

    row = 1
    ws.cell(row=row, column=1, value="月度统计").font = title_font
    row += 1
    headers = ["月份", "发票数量", "价税合计(元)"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = cell_border
    row += 1
    for k in sorted(month_map.keys()):
        ws.cell(row=row, column=1, value=k).border = cell_border
        ws.cell(row=row, column=2, value=month_map[k]["count"]).border = cell_border
        ws.cell(row=row, column=3, value=round(month_map[k]["amount"], 2)).border = cell_border
        row += 1

    for col in range(1, 4):
        ws.column_dimensions[get_column_letter(col)].width = 18


def _write_seller_stats_sheet(ws, records, header_fill, header_font, cell_border, title_font):
    from openpyxl.styles import Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    seller_map = defaultdict(lambda: {"count": 0, "amount": 0.0})

    for rec in records:
        seller = rec.get("销售方名称", "未知")
        amount = _parse_amount(rec.get("价税合计", ""))
        seller_map[seller]["count"] += 1
        seller_map[seller]["amount"] += amount

    sorted_sellers = sorted(seller_map.items(), key=lambda x: x[1]["amount"], reverse=True)

    ws.cell(row=1, column=1, value="销售方分析").font = title_font
    headers = ["销售方名称", "发票数量", "价税合计(元)", "平均金额(元)"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = cell_border

    for row_idx, (seller, data) in enumerate(sorted_sellers, 3):
        ws.cell(row=row_idx, column=1, value=seller).border = cell_border
        ws.cell(row=row_idx, column=2, value=data["count"]).border = cell_border
        ws.cell(row=row_idx, column=3, value=round(data["amount"], 2)).border = cell_border
        avg = round(data["amount"] / data["count"], 2) if data["count"] > 0 else 0
        ws.cell(row=row_idx, column=4, value=avg).border = cell_border

    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 20


def _write_reimb_stats_sheet(ws, records, header_fill, header_font, cell_border, title_font):
    from openpyxl.styles import Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    done_total = 0.0
    pending_total = 0.0
    done_count = 0
    pending_count = 0
    normal_count = 0
    cancel_count = 0
    red_count = 0

    for rec in records:
        amount = _parse_amount(rec.get("价税合计", ""))
        status = rec.get("报销状态", "未报销")
        invoice_status = rec.get("发票状态", "正常")

        if status == "已报销":
            done_total += amount
            done_count += 1
        else:
            pending_total += amount
            pending_count += 1

        if invoice_status == "正常":
            normal_count += 1
        elif invoice_status == "作废":
            cancel_count += 1
        elif invoice_status == "红冲":
            red_count += 1

    ws.cell(row=1, column=1, value="报销汇总").font = title_font

    headers = ["状态", "发票数量", "价税合计(元)"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = cell_border

    ws.cell(row=3, column=1, value="已报销").border = cell_border
    ws.cell(row=3, column=2, value=done_count).border = cell_border
    ws.cell(row=3, column=3, value=round(done_total, 2)).border = cell_border

    ws.cell(row=4, column=1, value="未报销").border = cell_border
    ws.cell(row=4, column=2, value=pending_count).border = cell_border
    ws.cell(row=4, column=3, value=round(pending_total, 2)).border = cell_border

    ws.cell(row=5, column=1, value="合计").border = cell_border
    ws.cell(row=5, column=1).font = Font(bold=True)
    ws.cell(row=5, column=2, value=done_count + pending_count).border = cell_border
    ws.cell(row=5, column=3, value=round(done_total + pending_total, 2)).border = cell_border

    row = 7
    ws.cell(row=row, column=1, value="发票状态").font = title_font
    row += 1
    headers2 = ["状态", "发票数量"]
    for col, h in enumerate(headers2, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = cell_border

    ws.cell(row=row + 1, column=1, value="正常").border = cell_border
    ws.cell(row=row + 1, column=2, value=normal_count).border = cell_border

    ws.cell(row=row + 2, column=1, value="作废").border = cell_border
    ws.cell(row=row + 2, column=2, value=cancel_count).border = cell_border

    ws.cell(row=row + 3, column=1, value="红冲").border = cell_border
    ws.cell(row=row + 3, column=2, value=red_count).border = cell_border

    for col in range(1, 4):
        ws.column_dimensions[get_column_letter(col)].width = 16


def _parse_amount(val):
    """解析金额字符串为浮点数。"""
    if not val:
        return 0.0
    try:
        return float(str(val).replace(",", "").replace("¥", "").replace("￥", ""))
    except (ValueError, TypeError):
        return 0.0


def _parse_date(date_str):
    """解析日期字符串，返回年月信息。"""
    import re
    m = re.match(r"(\d{4})年(\d{1,2})月(\d{1,2})日", date_str)
    if m:
        return {"year": int(m.group(1)), "month": int(m.group(2)), "day": int(m.group(3))}
    return None


if __name__ == "__main__":
    app.run(debug=False, port=5678, host="127.0.0.1")
