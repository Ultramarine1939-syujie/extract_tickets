"""导入既有 CSV / Excel 表格。"""

from __future__ import annotations

import csv
import io
import os
from collections.abc import Iterable

from .constants import DEFAULT_PENDING_REIMBURSEMENT, FIELDNAMES, FIELDNAMES_EXT, REIMBURSEMENT_FIELD

SUPPORTED_IMPORT_EXTENSIONS = {".csv", ".xlsx", ".xlsm"}


def import_table_bytes(file_bytes: bytes, filename: str) -> list[dict[str, str]]:
    """从 CSV 或 Excel 字节流读取票据记录。"""

    ext = os.path.splitext(filename.lower())[1]
    if ext == ".csv":
        return import_csv_bytes(file_bytes)
    if ext in {".xlsx", ".xlsm"}:
        return import_excel_bytes(file_bytes)
    raise ValueError("仅支持 CSV、XLSX 或 XLSM 表格")


def import_csv_bytes(file_bytes: bytes) -> list[dict[str, str]]:
    """读取导出的 CSV，兼容 UTF-8 BOM 和常见中文 Windows 编码。"""

    text = _decode_csv_text(file_bytes)
    reader = csv.DictReader(io.StringIO(text))
    return normalize_records(reader)


def import_excel_bytes(file_bytes: bytes) -> list[dict[str, str]]:
    """读取导出的 Excel，优先使用“行程明细”Sheet。"""

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("服务端未安装 openpyxl，请联系管理员") from exc

    workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    sheet = workbook["行程明细"] if "行程明细" in workbook.sheetnames else workbook.active
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    except StopIteration:
        return []

    raw_records = (
        {headers[index]: _cell_to_text(value) for index, value in enumerate(row) if index < len(headers)}
        for row in rows
    )
    return normalize_records(raw_records)


def normalize_records(records: Iterable[dict]) -> list[dict[str, str]]:
    """按项目标准字段清洗导入记录，忽略完全空行。"""

    normalized = []
    for raw in records:
        record = {field: _cell_to_text(raw.get(field, "")) for field in FIELDNAMES_EXT}
        if not any(record.get(field) for field in FIELDNAMES):
            continue
        if record["票价(元)"]:
            record["票价(元)"] = _format_price(record["票价(元)"])
        if not record[REIMBURSEMENT_FIELD]:
            record[REIMBURSEMENT_FIELD] = DEFAULT_PENDING_REIMBURSEMENT
        normalized.append(record)
    return normalized


def _cell_to_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _decode_csv_text(file_bytes: bytes) -> str:
    for encoding in ("utf-8-sig", "gb18030"):
        try:
            return file_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise UnicodeDecodeError("utf-8-sig", file_bytes, 0, len(file_bytes), "unsupported CSV encoding")


def _format_price(value: str) -> str:
    try:
        return f"{float(value):.2f}"
    except ValueError:
        return value
